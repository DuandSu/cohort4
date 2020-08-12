import datetime
import xls_app_tools
from openpyxl import load_workbook

def test_canAssertTrue():
    assert True

def test_TDD_Pipes():
    printLog = []
    printLog.append("Testing the TDD pipes")
    assert printLog[0] == "Testing the TDD pipes"

def test_play_with_dates():
    print("Playing with Dates:")
    playDate = datetime.datetime(2020, 7, 20, 0, 0)
    print(playDate.strftime("%Y-%m-%d"))
    playDStr = "2020-07-20 00:00"[0:10]
    print(playDStr)

def test_excel_date():
    import xlrd #ModuleNotFoundError: No module named 'xlrd'
    wb = load_workbook(filename = 'MRGTesting.xlsx', read_only=True)
    ws = wb['Invoice']

    for row in ws.iter_rows(values_only=True):
        justaVar = "No Date Set Yet"
        # justaVar = row[1] #prints 2020-07-20 00:00:00
        # justaVar = row[1][0:10] #TypeError: 'datetime.datetime' object is not subscriptable
        # justaVar = row[1].strftime("%Y-%m-%d") #AttributeError: 'str' object has no attribute 'strftime'.
        # justaVar = row[1].format()#AttributeError: 'datetime.datetime' object has no attribute 'format'
        # justaVar = row[1].date() #AttributeError: 'str' object has no attribute 'date'
        # justaVar = row[1].isoformat() #AttributeError: 'str' object has no attribute 'isoformat'
        # justaVar = datetime(*xlrd.xldate_as_tuple(row[1], 0))# TypeError: '<' not supported between instances of 'str' and 'float'
        # justaVar = ToDate(row[1]) #NameError: name 'ToDate' is not defined
        print(justaVar)
        # year = justaVar.year #AttributeError: 'str' object has no attribute 'year'
        # print(year)

def test_loadLocation():
    locDict = xls_app_tools.loadLocation("MRGTesting.xlsx")
    locDict.pop("locType")
    assert locDict == {1: "Head", 2: "Branch"}
    # header, *data = locDict
    # print(header.value)
    # print(data)
    # assert header["locType"] == "locDesc"
    # assert header == {"locType": "locDesc"}
    # assert locDict.difference({"locType": "locDesc"})

def test_loadCompany():
    compDict = xls_app_tools.loadCompany("MRGTesting.xlsx")
    assert compDict == {
            "compCode": "compName", 
            1: "Safeway",
            2: "Sobys",
            3: "Calgary Co-op",
            4: "Canadian Superstore",
            5: "Costco Wholesale"
        }

def test_loadCustomer():
    custDict = xls_app_tools.loadCustomer("MRGTesting.xlsx")
    assert custDict == {
            "locCode": ('compCode', 'cName', 'cLocType', 'cLocAddrKISS', 'cLocPhone', 'cLocEmail'), 
            1: (1, 'Head Office', 1, '181 Safeway Road, Toronto, ON', None, None),
            2: (1, 'Shawnessy', 2, '199 Shawnessey Drive, Calgary, AB', None, None),
            3: (1, 'Chestermere', 2, '199 Chestermer Road, Chestermere, AB', None, None),
            4: (1, 'Okotoks', 2, '199 Main Street, Okotoks, AB', None, None),
            5: (5, 'Canada Hub', 1, '202 Premier Avenue, Ottawa, ON', None, None),
            6: (5, 'Okotoks', 2, '12 River Street, Okotoks, AB', None, None),
            7: (5, 'Deer Meadows', 2, '800 Heritage Drive, Calgary, AB', None, None),
            8: (2, 'Bridlewood', 2, '111 Bridle Avenue, Calgary, AB', None, None),
            9: (2, 'Main Centre', 1, '888 Stephen Avenue, Calgary, AB', None, None),
            10: (3, 'Calgary Centre', 1, '404 3 Avenue, Calgary, AB', None, None),
            11: (3, 'Shawnessy', 2, '209 Shawnessey Drive, Calgary, AB', None, None),
            12: (4, 'Ottawa Main', 1, '101 Advocate Drive, Ottawa, ON', None, None),
            13: (4, 'Shawnessy', 2, '100 Shawnessey Drive, Calgary, AB', None, None)
        }

def test_loadProduct():
    prodDict = xls_app_tools.loadProduct("MRGTesting.xlsx")
    assert prodDict == {
            "prodNo": ('prodName', 'unitPrice', 'unitType'),
            1: ('Carrots', 2.99, 'Pkg'),
            2: ('Peas', 1.99, 'Lb'),
            3: ('Green Beans', 2.99, 'Lb'),
            4: ('Corn', 0.3, 'Each'),
            5: ('Cabbage', 3.99, 'Each'),
            6: ('Lettuce', 3.59, 'Each'),
            7: ('Brocolli', 2.59, 'Lb'),
            8: ('Califlower', 2.99, 'Lb'),
            9: ('Snap Peas', 2.99, 'Lb'),
            10: ('Spinach', 2.59, 'Lb')
        }

def test_loadInvoice():
    invDict = xls_app_tools.loadInvoice("MRGTesting.xlsx")
    assert invDict == {
            'invNo': ('invDate', 'custNo'),
            1: (datetime.datetime(2020, 7, 20, 0, 0), 2),
            2: (datetime.datetime(2020, 7, 19, 0, 0), 2)
        }

def test_loadInvLine():
    invLnDict = xls_app_tools.loadInvLine("MRGTesting.xlsx")
    assert  invLnDict == {
        ('invNo', 'lineNo') : ('prodNo', 'prodQty', 'prodPrice', 'prodUnit', 'taxCode', 'taxCodeRate'),
        (1, 1) : (1, 1, 2.99, 'Pkg', None, None),
        (1, 2) : (2, 2, 1.99, 'Lb', None, None),
        (1, 3) : (6, 3, 3.59, 'Each', None, None),
        (1, 4) : (8, 4, 2.99, 'Lb', None, None),
        (2, 1) : (1, 2, 1.99, 'Pkg', None, None),
        (2, 2) : (2, 3, 0.99, 'Lb', None, None),
        (2, 3) : (6, 4, 2.59, 'Each', None, None),
        (2, 4) : (8, 5, 1.99, 'Lb', None, None)
    }

    invLnDict.pop(('invNo', 'lineNo'))
    assert  invLnDict == {
        (1, 1) : (1, 1, 2.99, 'Pkg', None, None),
        (1, 2) : (2, 2, 1.99, 'Lb', None, None),
        (1, 3) : (6, 3, 3.59, 'Each', None, None),
        (1, 4) : (8, 4, 2.99, 'Lb', None, None),
        (2, 1) : (1, 2, 1.99, 'Pkg', None, None),
        (2, 2) : (2, 3, 0.99, 'Lb', None, None),
        (2, 3) : (6, 4, 2.59, 'Each', None, None),
        (2, 4) : (8, 5, 1.99, 'Lb', None, None)
    }
    invLnDict2JSON = {str(invLi[0]) + "/" + str(invLi[1]): invLnDict[invLi] for invLi in invLnDict}
    assert invLnDict2JSON == {
        '1/1' : (1, 1, 2.99, 'Pkg', None, None),
        '1/2' : (2, 2, 1.99, 'Lb', None, None),
        '1/3' : (6, 3, 3.59, 'Each', None, None),
        '1/4' : (8, 4, 2.99, 'Lb', None, None),
        '2/1' : (1, 2, 1.99, 'Pkg', None, None),
        '2/2' : (2, 3, 0.99, 'Lb', None, None),
        '2/3' : (6, 4, 2.59, 'Each', None, None),
        '2/4' : (8, 5, 1.99, 'Lb', None, None)
    }

def test_getInv():
    invDict = {
        'invNo': ('invDate', 'custNo'),
        1: (datetime.datetime(2020, 7, 20, 0, 0), 2),
        2: (datetime.datetime(2020, 7, 19, 0, 0), 2)
    }
    invLnDict = {
        'invNo': ('lineNo', 'prodNo', 'prodQty', 'prodPrice', 'prodUnit', 'taxCode', 'taxCodeRate'),
        (1, 1): (1, 2, 2.99, 'Pkg', None, None),
        (1, 2): (2, 2, 1.99, 'Lb', None, None),
        (1, 3): (6, 3, 3.59, 'Each', None, None),
        (1, 4): (8, 4, 2.99, 'Lb', None, None),
        (2, 1): (1, 2, 1.99, 'Pkg', None, None),
        (2, 2): (2, 3, 0.99, 'Lb', None, None),
        (2, 3): (6, 4, 2.59, 'Each', None, None),
        (2, 4): (8, 5, 1.99, 'Lb', None, None)
    }
    invNo = 1
    invToPrint = xls_app_tools.getInv(invNo, invDict, invLnDict)

    assert invToPrint == [
        (1, 1, datetime.datetime(2020, 7, 20, 0, 0), 2, 1, 2, 2.99, 'Pkg', None, None),
        (1, 2, datetime.datetime(2020, 7, 20, 0, 0), 2, 2, 2, 1.99, 'Lb', None, None),
        (1, 3, datetime.datetime(2020, 7, 20, 0, 0), 2, 6, 3, 3.59, 'Each', None, None),
        (1, 4, datetime.datetime(2020, 7, 20, 0, 0), 2, 8, 4, 2.99, 'Lb', None, None)
    ]
