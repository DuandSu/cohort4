import datetime
from openpyxl import load_workbook

def loadLocation(excelFile):
    wb = load_workbook(filename = excelFile, read_only=True)
    ws = wb['Location']

    locDict = {}
    for row in ws.iter_rows(values_only=True):
        locDict[row[0]] = row[1]

    return locDict

def loadCompany(excelFile):
    wb = load_workbook(filename = excelFile, read_only=True)
    ws = wb['Company']

    compDict = {}
    for row in ws.iter_rows(values_only=True):
        compDict[row[0]] = row[1]

    return compDict

def loadCustomer(excelFile):
    wb = load_workbook(filename = excelFile, read_only=True)
    ws = wb['Customer']

    custDict = {}
    for row in ws.iter_rows(values_only=True):
        custDict[row[0]] = (row[1],row[2],row[3],row[4],row[5],row[6])

    return custDict

def loadProduct(excelFile):
    wb = load_workbook(filename = excelFile, read_only=True)
    ws = wb['Product']

    prodDict = {}
    for row in ws.iter_rows(values_only=True):
        prodDict[row[0]] = (row[1],row[2],row[3])

    return prodDict

def loadInvoice(excelFile):
    wb = load_workbook(filename = excelFile, read_only=True)
    ws = wb['Invoice']

    invDict = {}
    for row in ws.iter_rows(values_only=True):
        # yearComma = row[1].index(",") - 4
        # yearEnd = yearComma + 4
        # year = row[1][yearComma:yearEnd]
        # dStr = row[1].strftime("%Y-%m-%d")
        # dStr = row[1][0:10]
        invDict[row[0]] = (row[1], row[2])

    return invDict

def loadInvLine(excelFile):
    wb = load_workbook(filename = excelFile, read_only=True)
    ws = wb['InvoiceLine']

    invLnDict = {}

    for row in ws.iter_rows(values_only=True):
        invLnDict[(row[0], row[1])] = (row[2],row[3],row[4],row[5],row[6],row[7])

    return invLnDict

def getInv(invNo, invDict, invLnDict):

    invToPrint = []

    for inv in invDict:
        if inv == invNo:
            invDate = invDict[inv][0]
            custNo = invDict[inv][1]

    for invLi in invLnDict:
        if invLi[0] == invNo:
            # print(invLi, invLnDict[invLi])
            invToPrint.append((invLi[0], invLi[1], invDate, custNo, 
                    invLnDict[invLi][0], invLnDict[invLi][1], invLnDict[invLi][2],
                    invLnDict[invLi][3], invLnDict[invLi][4], invLnDict[invLi][5]
                ))

    return invToPrint

def loadXLS():
    # Load in Excel File
    xlsDict = []
    xlsDict.append(loadLocation("MunRobinsonGardens.xlsx"))
    xlsDict.append(loadCompany("MunRobinsonGardens.xlsx"))
    xlsDict.append(loadCustomer("MunRobinsonGardens.xlsx"))
    xlsDict.append(loadProduct("MunRobinsonGardens.xlsx"))
    xlsDict.append(loadInvoice("MunRobinsonGardens.xlsx"))
    xlsDict.append(loadInvLine("MunRobinsonGardens.xlsx"))

    return xlsDict

def print_invoice(invNo, disp, xlsDict):
    # Load in Excel File
    locDict = xlsDict[0]
    compDict = xlsDict[1]
    custDict = xlsDict[2]
    prodDict = xlsDict[3]
    invDict = xlsDict[4]
    invLnDict = xlsDict[5]

    Column_Width = 80

    # Get Invoice Details
    # invNo = 1
    # invNo = 2
    invToPrint = getInv(invNo, invDict, invLnDict)
    # Prepare for print invoice header section by grabbing data, lookups for description and 
    # formatting. 
    invDate = invToPrint[0][2].strftime("%Y-%m-%d")
    invSep = 29 * " "
    invCustNo = invToPrint[0][3]
    invCustDesc = custDict[invCustNo][1]
    invCompNo = custDict[invCustNo][0]
    invCompDesc = compDict[invCompNo]

    custStaticWidth = 3 + 10 + 1 + 3
    custDescWidth = len(invCustDesc)
    compDescWidth = len(invCompDesc)
    custSepWidth = Column_Width - custStaticWidth - custDescWidth - compDescWidth
    custSep = custSepWidth * " "

    custAddr = custDict[invCustNo][3]

    addrStaticWidth = 12 + 3
    custAddrWidth = len(custAddr)
    addrSepWidth = Column_Width - addrStaticWidth - custAddrWidth
    addrSep = addrSepWidth * " "
    
    # If file, open and write to the file, otherwise print to terminal.
    if disp != "":
        f = open(disp, "w")
        f.write('')
        f.write(f'--------------------------------------------------------------------------------\r')
        f.write(f'-- Customer: {invCompDesc} {invCustDesc} {custSep}--\r')
        f.write(f'-- Invoice No: {"{:0>6d}".format(invNo)}    Invoice Date: {invDate}{invSep}--\r')
        f.write(f'-- Address: {custAddr}{addrSep} --\r')
        f.write(f'--------------------------------------------------------------------------------\r')
    else:
        print('')
        print(f'--------------------------------------------------------------------------------')
        print(f'-- Customer: {invCompDesc} {invCustDesc} {custSep}--')
        print(f'-- Invoice No: {"{:0>6d}".format(invNo)}    Invoice Date: {invDate}{invSep}--')
        print(f'-- Address: {custAddr}{addrSep} --')
        print(f'--------------------------------------------------------------------------------')

    # print invoice details.
    invTotal = 0
    for invLine in invToPrint:
        prodNo = invLine[4]
        prodDesc = prodDict[prodNo][0]
        prodQty = int(invLine[5])
        prodPrice = float(invLine[6])
        prodPriceDisp = "${:,.2f}".format(prodPrice)
        prodUnit = invLine[7]
        lineTot = prodQty * prodPrice
        invTotal += lineTot
        lineTotDisp = "${:,.2f}".format(lineTot)
        
        lineStaticWidth = 6 + 1 + 5 + 1 + 3
        prodDescWidth = len(prodDesc)
        prodQtyWidth = len(str(prodQty))
        prodPriceDispWidth = len(prodPriceDisp)
        prodUnitWidth = len(prodUnit)
        prodLineTotWidth = len(lineTotDisp)
        addrSepWidth = Column_Width - lineStaticWidth - prodDescWidth - prodQtyWidth - prodPriceDispWidth - prodUnitWidth - prodLineTotWidth
        lineSep = addrSepWidth * " "
        if disp != "":
            f.write(f'--    {prodDesc} {prodQty} for {prodPriceDisp}/{prodUnit}{lineSep}{lineTotDisp} --\r')
        else:
            print(f'--    {prodDesc} {prodQty} for {prodPriceDisp}/{prodUnit}{lineSep}{lineTotDisp} --')

    # print invoice summary
    invTotDisp = "${:,.2f}".format(invTotal)
    invTotStaticWidth = 6 + 6 + 3
    invTotDispWidth = len(invTotDisp)
    invTotSepWidth = Column_Width - invTotStaticWidth - invTotDispWidth
    invTotSep = invTotSepWidth * " "

    # If file, write to the file and close it, otherwise print to terminal.
    if disp != "":
        f.write(f'--                                                                  ------------\r')
        f.write(f'--    Total {invTotSep}{invTotDisp} --\r')
        f.write(f'--------------------------------------------------------------------------------\r')
        f.write(f'-- MunRobinson Gardens THANKS you for your business.                          --\r')
        f.write(f'-- Please Pay within 30 Days to avoid extra charges of 2 percent.             --\r') 
        f.write(f'--------------------------------------------------------------------------------\r')
        f.close()
    else:
        print(f'--                                                                  ------------')
        print(f'--    Total {invTotSep}{invTotDisp} --')
        print(f'--------------------------------------------------------------------------------')
        print(f'-- MunRobinson Gardens THANKS you for your business.                          --')
        print(f'-- Please Pay within 30 Days to avoid extra charges of 2 percent.             --') 
        print(f'--------------------------------------------------------------------------------')

